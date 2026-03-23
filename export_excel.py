import sqlite3
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# ----------------------------
# 1. Load Attendance from DB
# ----------------------------
conn = sqlite3.connect("students.db")
df = pd.read_sql_query("SELECT * FROM attendance", conn)
conn.close()

# ----------------------------
# 2. Prepare Excel file
# ----------------------------
file_path = "Attendance_Report_Final.xlsx"
sheet_name = "Attendance"

# if existing, load previous data
if os.path.exists(file_path):
    try:
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
    except Exception as e:
        print(f"[WARN] Could not open existing Excel: {e}")
        wb = load_workbook(file_path)
        ws = wb.active
else:
    # create new workbook and header row
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(["PRN", "Name"])  # base headers

# ----------------------------
# 3. Add new interval column
# ----------------------------
current_time = datetime.now().strftime("%H:%M")
interval_no = len(ws[1]) - 1  # exclude PRN/Name
interval_title = f"Interval {interval_no} ({current_time})"
ws.cell(row=1, column=len(ws[1]) + 1).value = interval_title

# ----------------------------
# 4. Update each student row
# ----------------------------
# Extract unique students from DB
students = df[["prn", "name"]].drop_duplicates().reset_index(drop=True)

# Build a dict for marking
attendance_dict = {str(row["prn"]): row for _, row in df.iterrows()}

# Prepare fills
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Ensure each student appears once
for idx, row in students.iterrows():
    prn = str(row["prn"])
    name = row["name"]

    # find or create row
    found = None
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value) == prn:
            found = r
            break
    if not found:
        ws.append([prn, name])
        found = ws.max_row

    # Determine presence
    if prn in attendance_dict:
        mark = "Present"
        ws.cell(row=found, column=len(ws[1])).value = f"Present ({attendance_dict[prn]['time']})"
        ws.cell(row=found, column=len(ws[1])).fill = green_fill
    else:
        ws.cell(row=found, column=len(ws[1])).value = "Absent"
        ws.cell(row=found, column=len(ws[1])).fill = red_fill

# ----------------------------
# 5. Save file
# ----------------------------
wb.save(file_path)
print(f"[INFO] Updated {file_path} with {interval_title}")

